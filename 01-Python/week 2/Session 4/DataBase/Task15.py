import openpyxl

class EmployeeManagementSystem:
    def __init__(self, file_name):
        """
        Initialize the EmployeeManagementSystem with the given Excel file name.

        :param file_name: The name of the Excel file to be created or loaded.
        """
        self.file_name = file_name
        self.workbook = self.create_excel_file()
        self.sheet = self.workbook["Employees"]

    def create_excel_file(self):
        """
        Create an Excel file with a predefined structure if it doesn't exist.

        :return: The workbook object.
        """
        try:
            # Try to load the existing workbook
            workbook = openpyxl.load_workbook(self.file_name)
        except FileNotFoundError:
            # Create a new workbook if the file doesn't exist
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Employees"
            # Add the header row
            sheet.append(["ID", "Name", "Job", "Salary"])
            # Save the workbook to the specified file
            workbook.save(self.file_name)
        return workbook

    def generate_id(self):
        """
        Generate a new unique ID for an employee.

        :return: A new unique ID string.
        """
        max_row = self.sheet.max_row
        if max_row == 1:
            return "IDX001"
        else:
            # Extract the numeric part of the last ID and increment it
            last_id = self.sheet.cell(row=max_row, column=1).value
            new_id_num = int(last_id[3:]) + 1
            new_id = f"IDX{new_id_num:03}"
            return new_id

    def add_employee(self, name, job, salary):
        """
        Add a new employee to the Excel sheet.

        :param name: The name of the employee.
        :param job: The job title of the employee.
        :param salary: The salary of the employee.
        """
        new_id = self.generate_id()
        self.sheet.append([new_id, name, job, salary])
        print(f"Employee '{name}' added successfully with ID {new_id}.")

    def print_employee_data(self):
        """
        Print the data of all employees from the Excel sheet.
        """
        if self.sheet.max_row == 1:
            print("No employees found.")
        else:
            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                print(f"ID: {row[0]}, Name: {row[1]}, Job: {row[2]}, Salary: {row[3]}")

    def remove_employee(self, employee_id):
        """
        Remove an employee from the Excel sheet by their ID.

        :param employee_id: The ID of the employee to be removed.
        """
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == employee_id:
                self.sheet.delete_rows(row[0].row)
                print(f"Employee with ID '{employee_id}' removed successfully.")
                return
        print(f"No employee found with ID '{employee_id}'.")

    def update_employee(self, employee_id, name=None, job=None, salary=None):
        """
        Update an employee's data in the Excel sheet.

        :param employee_id: The ID of the employee to be updated.
        :param name: The new name of the employee (optional).
        :param job: The new job title of the employee (optional).
        :param salary: The new salary of the employee (optional).
        """
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == employee_id:
                if name:
                    self.sheet.cell(row=row[0].row, column=2, value=name)
                if job:
                    self.sheet.cell(row=row[0].row, column=3, value=job)
                if salary:
                    self.sheet.cell(row=row[0].row, column=4, value=salary)
                print(f"Employee with ID '{employee_id}' updated successfully.")
                return
        print(f"No employee found with ID '{employee_id}'.")

    @staticmethod
    def display_menu():
        """
        Display the menu options for the Employee Management System.
        """
        print("\nEmployee Management System")
        print("1. Add new employee")
        print("2. Print employee data")
        print("3. Remove employee")
        print("4. Update employee")
        print("5. Exit")


def main():
    """
    Main function to run the Employee Management System.
    """
    file_name = "employees.xlsx"
    ems = EmployeeManagementSystem(file_name)

    while True:
        # Display the menu and get user input
        EmployeeManagementSystem.display_menu()
        choice = input("Enter your choice: ")

        if choice == '1':
            # Add new employee
            name = input("Enter employee name: ")
            job = input("Enter employee job: ")
            salary = float(input("Enter employee salary: "))
            ems.add_employee(name, job, salary)

        elif choice == '2':
            # Print employee data
            ems.print_employee_data()

        elif choice == '3':
            # Remove an employee
            employee_id = input("Enter employee ID to remove: ")
            ems.remove_employee(employee_id)

        elif choice == '4':
            # Update an employee's data
            employee_id = input("Enter employee ID to update: ")
            name = input("Enter new name (leave blank to skip): ")
            job = input("Enter new job (leave blank to skip): ")
            salary = input("Enter new salary (leave blank to skip): ")
            ems.update_employee(employee_id, name=name, job=job, salary=float(salary) if salary else None)

        elif choice == '5':
            # Exit the system
            print("Exiting the system.")
            break

        else:
            print("Invalid choice. Please try again.")

        # Save the changes to the workbook
        ems.workbook.save(file_name)


if __name__ == "__main__":
    main()
